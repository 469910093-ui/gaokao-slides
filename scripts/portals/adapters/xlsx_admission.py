"""XLSX 投档表解析。"""

from __future__ import annotations

import io
from typing import Any

from openpyxl import load_workbook

from admission_filter_lib import primary_undergrad_batch, parse_min_score
from portals.types import AdmissionRow, Artifact, ParseResult


def parse_xlsx_admission(artifact: Artifact, data: bytes) -> ParseResult:
    batch_name = artifact.batch or primary_undergrad_batch(artifact.province)
    track = artifact.track or "综合类"
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    rows: list[AdmissionRow] = []
    for ws in wb.worksheets:
        header: list[str] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            cells = [str(c).strip() if c is not None else "" for c in row]
            if i == 0:
                header = [c.lower() for c in cells]
                continue
            if not any(cells):
                continue
            school = _pick(cells, header, ("院校", "学校", "院校名称", "school"))
            score_raw = _pick(cells, header, ("总分", "投档分", "最低分", "min", "分数"))
            if not school:
                # 无表头：尝试固定列
                if len(cells) >= 6 and not cells[2].isdigit():
                    school = cells[2]
                    score_raw = cells[5]
            score = parse_min_score(score_raw)
            if not school or score is None:
                continue
            grp = _pick(cells, header, ("专业组", "组号", "group")) or (cells[3] if len(cells) > 3 else "")
            info = _pick(cells, header, ("选考", "选科", "科目")) or (cells[4] if len(cells) > 4 else "")
            rows.append(
                AdmissionRow(
                    province=artifact.province,
                    year=artifact.year,
                    track=track,
                    schoolName=school,
                    minScore=score,
                    batch=batch_name,
                    groupName=str(grp),
                    groupInfo=str(info),
                    sourceUrl=artifact.url,
                )
            )
    return ParseResult(artifact=artifact, rows=rows, parser="xlsx_openpyxl")


def _pick(cells: list[str], header: list[str], keys: tuple[str, ...]) -> str:
    for key in keys:
        for idx, h in enumerate(header):
            if key in h and idx < len(cells):
                return cells[idx]
    return ""
