"""河北省教育考试院本科批平行志愿投档 XLSX（按专业公布投档最低分）。"""

from __future__ import annotations

import io
import re

from openpyxl import load_workbook

from admission_filter_lib import primary_undergrad_batch, parse_min_score
from portals.types import AdmissionRow, Artifact, ParseResult

_SCHOOL_RE = re.compile(r"^(.+?)(?:\([^)]+\))?(?:\[[^\]]+\])?$")


def _clean_school(raw: str) -> str:
    text = (raw or "").replace("\n", "").strip()
    m = _SCHOOL_RE.match(text)
    return (m.group(1) if m else text).strip()


def parse_xlsx_hebei_admission(artifact: Artifact, data: bytes) -> ParseResult:
    batch_name = artifact.batch or primary_undergrad_batch(artifact.province)
    track = artifact.track or "物理类"
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows: list[AdmissionRow] = []
    started = False

    for row in ws.iter_rows(values_only=True):
        cells = [c for c in row]
        if not any(c is not None and str(c).strip() for c in cells):
            continue
        code = str(cells[0]).strip().replace(".0", "") if cells[0] is not None else ""
        if not started:
            if code.isdigit() and len(code) >= 4:
                started = True
            else:
                continue
        if len(cells) < 5:
            continue
        school_raw = str(cells[1] or "").strip()
        major = str(cells[3] or "").strip()
        score = parse_min_score(cells[4])
        school = _clean_school(school_raw)
        if not school or score is None:
            continue
        rows.append(
            AdmissionRow(
                province=artifact.province,
                year=artifact.year,
                track=track,
                schoolName=school,
                minScore=score,
                batch=batch_name,
                groupName=major[:40] if major else "",
                groupInfo=major,
                schoolCode=code[:4],
                sourceUrl=artifact.url,
            )
        )

    return ParseResult(artifact=artifact, rows=rows, parser="xlsx_hebei")
