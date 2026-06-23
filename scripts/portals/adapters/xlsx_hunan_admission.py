"""湖南省本科批平行志愿投档 XLSX（院校专业组）。"""

from __future__ import annotations

import io

from openpyxl import load_workbook

from admission_filter_lib import parse_min_score, primary_undergrad_batch
from portals.types import AdmissionRow, Artifact, ParseResult


def parse_xlsx_hunan_admission(artifact: Artifact, data: bytes) -> ParseResult:
    batch_name = artifact.batch or primary_undergrad_batch(artifact.province)
    default_track = artifact.track or "物理类"
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows: list[AdmissionRow] = []
    header_seen = False

    for row in ws.iter_rows(values_only=True):
        cells = list(row)
        if not any(c is not None and str(c).strip() for c in cells):
            continue
        first = str(cells[0] or "").strip()
        if not header_seen:
            if first == "批次" or (len(cells) > 7 and str(cells[7] or "") == "投档线"):
                header_seen = True
            continue
        if len(cells) < 8:
            continue
        track_raw = str(cells[2] or "")
        if "历史" in track_raw:
            track = "历史类"
        elif "物理" in track_raw:
            track = "物理类"
        else:
            track = default_track
        if artifact.track and artifact.track not in ("综合类", "") and track != artifact.track:
            continue
        code = str(cells[3] or "").strip().replace(".0", "")
        school = str(cells[4] or "").strip()
        group_no = str(cells[5] or "").strip().replace(".0", "")
        group_name = str(cells[6] or "").strip()
        score = parse_min_score(cells[7])
        if not code.isdigit() or not school or score is None:
            continue
        note = str(cells[15] or "").strip() if len(cells) > 15 else ""
        rows.append(
            AdmissionRow(
                province=artifact.province,
                year=artifact.year,
                track=track,
                schoolName=school,
                minScore=score,
                batch=batch_name,
                groupName=f"第{group_no}组" if group_no else group_name,
                groupInfo=group_name or note,
                schoolCode=code,
                sourceUrl=artifact.url,
            )
        )

    return ParseResult(artifact=artifact, rows=rows, parser="xlsx_hunan")
